# Standard Libary
import logging

# Third-Party
from openpyxl import Workbook

# Django
from django.apps import apps as api_apps

# First-Party

config = api_apps.get_app_config('api')

log = logging.getLogger(__name__)


def create_bbscores_excel(session):
    Entry = config.get_model('Entry')
    wb = Workbook()
    ws = wb.active
    fieldnames = [
        'oa',
        'contestant_id',
        'group_name',
        'group_type',
        'song_number',
        'song_title',
    ]
    ws.append(fieldnames)
    entries = session.entries.filter(
        status=Entry.STATUS.approved,
    ).order_by('draw')
    for entry in entries:
        oa = entry.draw
        group_name = entry.group.name.encode('utf-8').strip()
        group_type = entry.group.get_kind_display()
        if group_type == 'Quartet':
            contestant_id = entry.group.bhs_id
        elif group_type == 'Chorus':
            contestant_id = entry.group.code
        else:
            raise RuntimeError("Improper Entity Type")
        i = 1
        for repertory in entry.group.repertories.order_by('chart__title'):
            song_number = i
            song_title = repertory.chart.title.encode('utf-8').strip()
            i += 1
            row = [
                oa,
                contestant_id,
                group_name,
                group_type,
                song_number,
                song_title,
            ]
            ws.append(row)
    wb.save('bbscores.xlsx')


def create_drcj_report_excel(session):
    Entry = config.get_model('Entry')
    Group = config.get_model('Group')
    Participant = config.get_model('Participant')
    wb = Workbook()
    ws = wb.active
    fieldnames = [
        'oa',
        'group_name',
        'representing',
        'evaluation',
        'private',
        'bhs_id',
        'group_exp_date',
        'repertory_count',
        'particpant_count',
        'expiring_count',
        'tenor',
        'lead',
        'baritone',
        'bass',
        'directors',
        'awards',
        'chapters',
    ]
    ws.append(fieldnames)
    entries = session.entries.filter(
        status=Entry.STATUS.approved,
    ).order_by('draw')
    for entry in entries:
        oa = entry.draw
        group_name = entry.group.name
        representing = entry.group.organization.name
        evaluation = entry.is_evaluation
        private = entry.is_private
        bhs_id = entry.group.bhs_id
        repertory_count = entry.group.repertories.filter(
            status__gt=0,
        ).count()
        group_exp_date = None
        repertory_count = entry.group.repertories.filter(
            status__gt=0,
        ).count()
        participants = entry.participants.filter(
            status__gt=0,
        )
        participant_count = participants.count()
        expiring_count = participants.filter(
            person__current_through__lte=session.convention.close_date,
        ).count()
        directors = entry.directors
        awards_list = []
        contestants = entry.contestants.filter(
            status__gt=0,
        ).order_by('contest__award__name')
        for contestant in contestants:
            awards_list.append(contestant.contest.award.name)
        awards = "\n".join(filter(None, awards_list))
        parts = {}
        part = 1
        while part <= 4:
            try:
                participant = entry.participants.get(
                    part=part,
                )
            except Participant.DoesNotExist:
                parts[part] = None
                part += 1
                continue
            except Participant.MultipleObjectsReturned:
                parts[part] = None
                part += 1
                continue
            participant_list = []
            participant_list.append(
                participant.person.nomen,
            )
            participant_list.append(
                participant.person.email,
            )
            participant_list.append(
                participant.person.phone,
            )
            participant_detail = "\n".join(filter(None, participant_list))
            parts[part] = participant_detail
            part += 1
        chapters_list = []
        if entry.group.kind == entry.group.KIND.quartet:
            participants = entry.participants.filter(status__gt=0)
            for participant in participants:
                person_chapter_list = []
                for member in participant.person.members.filter(
                    status=10,
                    group__kind=Group.KIND.chorus,
                ).distinct('group'):
                    person_chapter_list.append(
                        member.group.name,
                    )
                chapters_list.extend(
                    person_chapter_list
                )
            dedupe = list(set(chapters_list))
            chapters = "\n".join(filter(None, dedupe))
        else:
            chapters = None
        row = [
            oa,
            group_name,
            representing,
            evaluation,
            private,
            bhs_id,
            group_exp_date,
            repertory_count,
            participant_count,
            expiring_count,
            parts[1],
            parts[2],
            parts[3],
            parts[4],
            directors,
            awards,
            chapters,
        ]
        ws.append(row)
    wb.save('drcj_report.xlsx')


def create_admin_emails_excel(session):
    Entry = config.get_model('Entry')
    wb = Workbook()
    ws = wb.active
    fieldnames = [
        'group',
        'admin',
        'email',
        'cell',
    ]
    ws.append(fieldnames)
    entries = session.entries.filter(
        status=Entry.STATUS.approved,
    ).order_by('draw')
    for entry in entries:
        admins = entry.group.members.filter(
            is_admin=True,
        )
        for admin in admins:
            group = entry.group.nomen.encode('utf-8').strip()
            person = admin.person.nomen.encode('utf-8').strip()
            email = admin.person.email.encode('utf-8').strip()
            cell = admin.person.cell_phone
            row = [
                group,
                person,
                email,
                cell,
            ]
            ws.append(row)
    wb.save('admin_emails.xlsx')
